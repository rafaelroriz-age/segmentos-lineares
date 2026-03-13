"""
Exportação robusta de resultados em Excel (.xlsx) com múltiplas abas,
gráficos embutidos, formatação condicional, cabeçalhos estilizados e
resumos estatísticos.

Usa openpyxl para gráficos nativos do Excel (abrem em qualquer leitor).
"""

import io
import numpy as np
import pandas as pd
from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import (
    BarChart, LineChart, Reference, BarChart3D, ScatterChart, Series
)
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule


# ── Estilos globais ───────────────────────────────────────────────────────

FONT_HEADER = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
FONT_TITLE  = Font(name='Calibri', bold=True, size=14, color='1F4E79')
FONT_SUB    = Font(name='Calibri', bold=True, size=12, color='2E75B6')
FONT_BODY   = Font(name='Calibri', size=10)
FONT_SMALL  = Font(name='Calibri', size=9, italic=True, color='808080')

FILL_HEADER = PatternFill(start_color='2E75B6', end_color='2E75B6',
                          fill_type='solid')
FILL_ALT    = PatternFill(start_color='D6E4F0', end_color='D6E4F0',
                          fill_type='solid')
FILL_GREEN  = PatternFill(start_color='C6EFCE', end_color='C6EFCE',
                          fill_type='solid')
FILL_YELLOW = PatternFill(start_color='FFEB9C', end_color='FFEB9C',
                          fill_type='solid')
FILL_RED    = PatternFill(start_color='FFC7CE', end_color='FFC7CE',
                          fill_type='solid')

ALIGN_CENTER = Alignment(horizontal='center', vertical='center',
                         wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal='left', vertical='center',
                         wrap_text=True)

BORDER_THIN = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0'),
)


def _write_df(ws, df, start_row=1, start_col=1, with_header=True,
              alternating=True):
    """Escreve DataFrame em worksheet com formatação."""
    rows = list(dataframe_to_rows(df, index=False, header=with_header))
    for r_idx, row in enumerate(rows):
        for c_idx, val in enumerate(row):
            cell = ws.cell(row=start_row + r_idx,
                           column=start_col + c_idx, value=val)
            cell.font = FONT_BODY
            cell.border = BORDER_THIN
            cell.alignment = ALIGN_CENTER

            if r_idx == 0 and with_header:
                cell.font = FONT_HEADER
                cell.fill = FILL_HEADER
                cell.alignment = ALIGN_CENTER
            elif alternating and r_idx % 2 == 0 and r_idx > 0:
                cell.fill = FILL_ALT

            # Formatar números
            if isinstance(val, float):
                cell.number_format = '#,##0.000'

    # Ajustar largura das colunas
    for c_idx in range(len(df.columns)):
        col_letter = get_column_letter(start_col + c_idx)
        max_len = max(
            len(str(df.columns[c_idx])),
            df.iloc[:, c_idx].astype(str).str.len().max() if len(df) > 0 else 0
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 30)

    return start_row + len(rows)


def _add_title(ws, title, row=1, col=1, merge_cols=6):
    """Adiciona título estilizado com merge."""
    ws.merge_cells(
        start_row=row, start_column=col,
        end_row=row, end_column=col + merge_cols - 1
    )
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = FONT_TITLE
    cell.alignment = ALIGN_LEFT
    return row + 2


def _add_subtitle(ws, title, row=1, col=1):
    """Adiciona subtítulo."""
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = FONT_SUB
    cell.alignment = ALIGN_LEFT
    return row + 1


def _add_bar_chart(ws, title, categories_col, values_cols, labels,
                   start_row, data_start_row, data_end_row,
                   anchor_cell='A1', width=18, height=12):
    """Adiciona gráfico de barras ao worksheet."""
    chart = BarChart()
    chart.type = 'col'
    chart.title = title
    chart.style = 10
    chart.width = width
    chart.height = height

    cats = Reference(ws, min_col=categories_col,
                     min_row=data_start_row + 1,
                     max_row=data_end_row)
    chart.set_categories(cats)

    for i, (col_idx, label) in enumerate(zip(values_cols, labels)):
        data = Reference(ws, min_col=col_idx,
                         min_row=data_start_row,
                         max_row=data_end_row)
        chart.add_data(data, titles_from_data=True)

    ws.add_chart(chart, anchor_cell)


def _add_line_chart(ws, title, cat_col, val_cols, labels,
                    data_start_row, data_end_row,
                    anchor_cell='A1', width=18, height=12):
    """Adiciona gráfico de linhas."""
    chart = LineChart()
    chart.title = title
    chart.style = 10
    chart.width = width
    chart.height = height

    cats = Reference(ws, min_col=cat_col,
                     min_row=data_start_row + 1,
                     max_row=data_end_row)
    chart.set_categories(cats)

    for col_idx in val_cols:
        data = Reference(ws, min_col=col_idx,
                         min_row=data_start_row,
                         max_row=data_end_row)
        chart.add_data(data, titles_from_data=True)

    ws.add_chart(chart, anchor_cell)


# ══════════════════════════════════════════════════════════════════════════════
# FUNÇÃO PRINCIPAL DE EXPORTAÇÃO
# ══════════════════════════════════════════════════════════════════════════════

def exportar_excel(
    seg120,
    resultados,
    df_comp,
    all_downloads,
    vars_ok,
    METODO_COL,
    df_diag=None,
    gap_df=None,
    k_otimo=None,
    ari_matrix=None,
    parametros=None,
):
    """
    Gera arquivo Excel completo com múltiplas abas:

    1. Resumo Geral       — parâmetros, data, estatísticas gerais
    2. Dados Agregados     — seg120 completo
    3. Comparação Métricas — tabela comparativa + gráficos
    4. Segmentos_{método}  — uma aba por método com segmentos finais
    5. Auditoria           — diagnóstico, Gap Statistic, ARI
    6. Perfil Longitudinal — dados para gráfico de perfil

    Retorna
    -------
    bytes : conteúdo do arquivo .xlsx pronto para download
    """
    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════
    # ABA 1 — RESUMO GERAL
    # ══════════════════════════════════════════════════════════════════════
    ws_resumo = wb.active
    ws_resumo.title = 'Resumo Geral'
    ws_resumo.sheet_properties.tabColor = '1F4E79'

    row = _add_title(ws_resumo, '🛣️ Segmentação Homogênea LINEAR — Relatório',
                     merge_cols=8)
    row = _add_subtitle(ws_resumo, 'Gerado automaticamente pelo pipeline', row)

    # Info geral
    info = {
        'Data de geração': datetime.now().strftime('%d/%m/%Y %H:%M'),
        'Total de segmentos (120m)': len(seg120),
        'Extensão total (m)': f"{seg120['km_fim'].max() - seg120['km_ini'].min():.0f}" if len(seg120) > 0 else 'N/A',
        'Deflexão média': f"{seg120['Deflexao'].mean():.2f}" if 'Deflexao' in seg120.columns else 'N/A',
        'Deflexão caract. média': f"{seg120['Deflexao_caract'].mean():.2f}" if 'Deflexao_caract' in seg120.columns else 'N/A',
        'Features usadas': ', '.join(vars_ok),
        'Métodos executados': ', '.join(resultados.keys()),
        'Total de métodos': len(resultados),
    }
    if parametros:
        info.update(parametros)

    for key, val in info.items():
        ws_resumo.cell(row=row, column=1, value=key).font = Font(bold=True, name='Calibri', size=10)
        ws_resumo.cell(row=row, column=1).border = BORDER_THIN
        ws_resumo.cell(row=row, column=2, value=str(val)).font = FONT_BODY
        ws_resumo.cell(row=row, column=2).border = BORDER_THIN
        ws_resumo.cell(row=row, column=2).alignment = ALIGN_LEFT
        row += 1

    ws_resumo.column_dimensions['A'].width = 28
    ws_resumo.column_dimensions['B'].width = 50

    # Tabela de métodos
    row += 1
    row = _add_subtitle(ws_resumo, 'Métodos Disponíveis', row)
    metodos_info = pd.DataFrame([
        {'Método': m, 'Tipo': t, 'Segmentos': len(set(r['labels']))}
        for m, r in resultados.items()
        for t in [('Tradicional' if m in ['CDA','MCV','SHS']
                    else 'Change-Point' if m == 'PELT'
                    else 'ML Supervisionado' if m in ['KNN','Random Forest']
                    else 'Clustering')]
    ])
    row = _write_df(ws_resumo, metodos_info, start_row=row)

    # ══════════════════════════════════════════════════════════════════════
    # ABA 2 — DADOS AGREGADOS
    # ══════════════════════════════════════════════════════════════════════
    ws_dados = wb.create_sheet('Dados Agregados')
    ws_dados.sheet_properties.tabColor = '548235'

    row = _add_title(ws_dados, 'Dados Agregados (segmentos)', merge_cols=len(seg120.columns))
    row = _write_df(ws_dados, seg120.round(4), start_row=row)

    # Gráfico de perfil de deflexão
    data_start = 3  # header row (row 3 = título row2 + header row)
    data_end = data_start + len(seg120)
    km_col = list(seg120.columns).index('km_ini') + 1
    defl_col = list(seg120.columns).index('Deflexao') + 1

    chart = LineChart()
    chart.title = 'Perfil Longitudinal de Deflexão'
    chart.style = 10
    chart.width = 24
    chart.height = 12
    chart.x_axis.title = 'Estação (m)'
    chart.y_axis.title = 'Deflexão (0,01mm)'

    cats = Reference(ws_dados, min_col=km_col, min_row=data_start + 1,
                     max_row=data_end)
    vals = Reference(ws_dados, min_col=defl_col, min_row=data_start,
                     max_row=data_end)
    chart.add_data(vals, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.line.width = 15000

    ws_dados.add_chart(chart, f'{get_column_letter(len(seg120.columns) + 2)}3')

    # Deflexão caract se existir
    if 'Deflexao_caract' in seg120.columns:
        dc_col = list(seg120.columns).index('Deflexao_caract') + 1
        vals2 = Reference(ws_dados, min_col=dc_col, min_row=data_start,
                          max_row=data_end)
        chart.add_data(vals2, titles_from_data=True)

    # ══════════════════════════════════════════════════════════════════════
    # ABA 3 — COMPARAÇÃO DE MÉTRICAS
    # ══════════════════════════════════════════════════════════════════════
    if df_comp is not None and not df_comp.empty:
        ws_comp = wb.create_sheet('Comparação Métricas')
        ws_comp.sheet_properties.tabColor = 'BF8F00'

        row = _add_title(ws_comp, 'Comparação de Métricas entre Métodos',
                         merge_cols=len(df_comp.columns))
        data_start_row = row
        row = _write_df(ws_comp, df_comp.round(4), start_row=row)
        data_end_row = row - 1

        # Formatação condicional: Silhouette (mais alto = melhor)
        if 'silhouette' in df_comp.columns:
            sil_col = list(df_comp.columns).index('silhouette') + 1
            sil_letter = get_column_letter(sil_col)
            ws_comp.conditional_formatting.add(
                f'{sil_letter}{data_start_row + 1}:{sil_letter}{data_end_row}',
                ColorScaleRule(
                    start_type='min', start_color='FFC7CE',
                    mid_type='percentile', mid_value=50, mid_color='FFEB9C',
                    end_type='max', end_color='C6EFCE'
                )
            )

        # Formatação condicional: Davies-Bouldin (mais baixo = melhor)
        if 'davies_bouldin' in df_comp.columns:
            db_col = list(df_comp.columns).index('davies_bouldin') + 1
            db_letter = get_column_letter(db_col)
            ws_comp.conditional_formatting.add(
                f'{db_letter}{data_start_row + 1}:{db_letter}{data_end_row}',
                ColorScaleRule(
                    start_type='min', start_color='C6EFCE',
                    mid_type='percentile', mid_value=50, mid_color='FFEB9C',
                    end_type='max', end_color='FFC7CE'
                )
            )

        # Gráfico de barras — Silhouette por método
        metodo_col_idx = 1  # primeira coluna = método
        chart_row = data_end_row + 2

        if 'silhouette' in df_comp.columns:
            chart1 = BarChart()
            chart1.type = 'col'
            chart1.title = 'Silhouette Score por Método'
            chart1.style = 10
            chart1.width = 18
            chart1.height = 10

            cats = Reference(ws_comp, min_col=metodo_col_idx,
                             min_row=data_start_row + 1,
                             max_row=data_end_row)
            chart1.set_categories(cats)

            data = Reference(ws_comp, min_col=sil_col,
                             min_row=data_start_row,
                             max_row=data_end_row)
            chart1.add_data(data, titles_from_data=True)
            chart1.series[0].graphicalProperties.solidFill = '2E75B6'

            dl = DataLabelList()
            dl.showVal = True
            dl.numFmt = '0.000'
            chart1.series[0].dLbls = dl

            ws_comp.add_chart(chart1, f'A{chart_row}')

        # Gráfico — Calinski-Harabász
        if 'calinski' in df_comp.columns:
            cal_col = list(df_comp.columns).index('calinski') + 1
            chart2 = BarChart()
            chart2.type = 'col'
            chart2.title = 'Calinski-Harabász por Método'
            chart2.style = 10
            chart2.width = 18
            chart2.height = 10

            chart2.set_categories(cats)
            data2 = Reference(ws_comp, min_col=cal_col,
                              min_row=data_start_row,
                              max_row=data_end_row)
            chart2.add_data(data2, titles_from_data=True)
            chart2.series[0].graphicalProperties.solidFill = '548235'

            dl2 = DataLabelList()
            dl2.showVal = True
            dl2.numFmt = '#,##0.0'
            chart2.series[0].dLbls = dl2

            ws_comp.add_chart(chart2, f'J{chart_row}')

        # Gráfico — Davies-Bouldin
        if 'davies_bouldin' in df_comp.columns:
            chart3 = BarChart()
            chart3.type = 'col'
            chart3.title = 'Davies-Bouldin por Método (menor = melhor)'
            chart3.style = 10
            chart3.width = 18
            chart3.height = 10

            chart3.set_categories(cats)
            data3 = Reference(ws_comp, min_col=db_col,
                              min_row=data_start_row,
                              max_row=data_end_row)
            chart3.add_data(data3, titles_from_data=True)
            chart3.series[0].graphicalProperties.solidFill = 'BF8F00'

            dl3 = DataLabelList()
            dl3.showVal = True
            dl3.numFmt = '0.000'
            chart3.series[0].dLbls = dl3

            ws_comp.add_chart(chart3, f'A{chart_row + 16}')

        # Gráfico agrupado — todas as métricas normalizadas
        # Normalizar para 0-1 para comparação visual
        numeric_cols = df_comp.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) >= 2:
            df_norm = df_comp.copy()
            for c in numeric_cols:
                rng = df_norm[c].max() - df_norm[c].min()
                if rng > 0:
                    df_norm[c] = (df_norm[c] - df_norm[c].min()) / rng
                else:
                    df_norm[c] = 0.5

            norm_start = data_end_row + 35
            ws_comp.cell(row=norm_start, column=1,
                         value='Métricas Normalizadas (0–1)').font = FONT_SUB
            norm_data_start = norm_start + 1
            norm_end = _write_df(ws_comp, df_norm.round(3),
                                 start_row=norm_data_start) - 1

            chart4 = BarChart()
            chart4.type = 'col'
            chart4.title = 'Métricas Normalizadas (0–1) — Comparação Visual'
            chart4.style = 10
            chart4.width = 24
            chart4.height = 12

            cats_norm = Reference(ws_comp, min_col=1,
                                  min_row=norm_data_start + 1,
                                  max_row=norm_end)
            chart4.set_categories(cats_norm)

            for col_idx in range(len(numeric_cols)):
                ref = Reference(ws_comp,
                                min_col=list(df_norm.columns).index(numeric_cols[col_idx]) + 1,
                                min_row=norm_data_start,
                                max_row=norm_end)
                chart4.add_data(ref, titles_from_data=True)

            ws_comp.add_chart(chart4, f'J{chart_row + 16}')

    # ══════════════════════════════════════════════════════════════════════
    # ABAS 4..N — SEGMENTOS POR MÉTODO
    # ══════════════════════════════════════════════════════════════════════
    for nome, (seg_final, tbl) in all_downloads.items():
        safe = nome.replace(' ', '_').replace('.', '').replace('+', '_')[:28]
        ws_met = wb.create_sheet(f'Seg_{safe}')
        ws_met.sheet_properties.tabColor = '2E75B6'

        col_cluster = METODO_COL.get(nome, 'cluster_kmeans')
        col_final = f'{col_cluster}_espacial'

        # Título
        row = _add_title(ws_met, f'Segmentos Finais — {nome}', merge_cols=10)

        # Info do método
        tipo = ('Tradicional' if nome in ['CDA', 'MCV', 'SHS']
                else 'Change-Point' if nome == 'PELT'
                else 'ML Supervisionado' if nome in ['KNN', 'Random Forest']
                else 'Clustering + Linearização')
        ws_met.cell(row=row, column=1, value='Tipo:').font = Font(bold=True, name='Calibri')
        ws_met.cell(row=row, column=2, value=tipo).font = FONT_BODY
        row += 1
        ws_met.cell(row=row, column=1, value='Segmentos:').font = Font(bold=True, name='Calibri')
        ws_met.cell(row=row, column=2, value=len(tbl)).font = FONT_BODY
        row += 1
        if 'Comprimento_m' in tbl.columns:
            ws_met.cell(row=row, column=1, value='Extensão total (m):').font = Font(bold=True, name='Calibri')
            ws_met.cell(row=row, column=2, value=f"{tbl['Comprimento_m'].sum():.0f}").font = FONT_BODY
            row += 1

        row += 1

        # Tabela resumo dos segmentos
        row = _add_subtitle(ws_met, 'Resumo dos Segmentos', row)
        tbl_start = row
        row = _write_df(ws_met, tbl.round(2), start_row=row)
        tbl_end = row - 1

        # Formatação condicional na deflexão
        if 'Deflexao_med' in tbl.columns:
            defl_idx = list(tbl.columns).index('Deflexao_med') + 1
            defl_letter = get_column_letter(defl_idx)
            ws_met.conditional_formatting.add(
                f'{defl_letter}{tbl_start + 1}:{defl_letter}{tbl_end}',
                ColorScaleRule(
                    start_type='min', start_color='C6EFCE',
                    mid_type='percentile', mid_value=50, mid_color='FFEB9C',
                    end_type='max', end_color='FFC7CE'
                )
            )

        # Gráfico de barras — Deflexão média por segmento
        if 'Deflexao_med' in tbl.columns and len(tbl) > 0:
            chart_seg = BarChart()
            chart_seg.type = 'col'
            chart_seg.title = f'Deflexão Média por Segmento — {nome}'
            chart_seg.style = 10
            chart_seg.width = 20
            chart_seg.height = 10
            chart_seg.y_axis.title = 'Deflexão (0,01mm)'

            cluster_col_idx = 1
            cats = Reference(ws_met, min_col=cluster_col_idx,
                             min_row=tbl_start + 1, max_row=tbl_end)
            chart_seg.set_categories(cats)

            data = Reference(ws_met, min_col=defl_idx,
                             min_row=tbl_start, max_row=tbl_end)
            chart_seg.add_data(data, titles_from_data=True)
            chart_seg.series[0].graphicalProperties.solidFill = '2E75B6'

            dl = DataLabelList()
            dl.showVal = True
            dl.numFmt = '0.00'
            chart_seg.series[0].dLbls = dl

            ws_met.add_chart(chart_seg, f'A{tbl_end + 2}')

        # Gráfico de comprimento dos segmentos
        if 'Comprimento_m' in tbl.columns and len(tbl) > 0:
            comp_idx = list(tbl.columns).index('Comprimento_m') + 1
            chart_comp = BarChart()
            chart_comp.type = 'col'
            chart_comp.title = f'Comprimento dos Segmentos — {nome}'
            chart_comp.style = 10
            chart_comp.width = 20
            chart_comp.height = 10
            chart_comp.y_axis.title = 'Comprimento (m)'

            chart_comp.set_categories(cats)
            data_c = Reference(ws_met, min_col=comp_idx,
                               min_row=tbl_start, max_row=tbl_end)
            chart_comp.add_data(data_c, titles_from_data=True)
            chart_comp.series[0].graphicalProperties.solidFill = '548235'

            dl_c = DataLabelList()
            dl_c.showVal = True
            dl_c.numFmt = '#,##0'
            chart_comp.series[0].dLbls = dl_c

            ws_met.add_chart(chart_comp, f'L{tbl_end + 2}')

        # Gráfico de perfil — deflexão ao longo da rodovia com cor por segmento
        if col_final in seg_final.columns and 'km_ini' in seg_final.columns:
            # Tabela auxiliar para perfil
            perf_start = tbl_end + 20
            perf_start = _add_subtitle(ws_met, 'Perfil Longitudinal', perf_start)

            perf_df = seg_final[['km_ini', 'Deflexao', col_final]].copy()
            perf_df = perf_df.rename(columns={col_final: 'Segmento'})
            perf_data_start = perf_start
            perf_end = _write_df(ws_met, perf_df.round(2),
                                 start_row=perf_start)
            perf_data_end = perf_end - 1

            chart_perf = LineChart()
            chart_perf.title = f'Perfil Longitudinal — {nome}'
            chart_perf.style = 10
            chart_perf.width = 24
            chart_perf.height = 12
            chart_perf.x_axis.title = 'Estação (m)'
            chart_perf.y_axis.title = 'Deflexão (0,01mm)'

            cats_p = Reference(ws_met, min_col=1,
                               min_row=perf_data_start + 1,
                               max_row=perf_data_end)
            vals_p = Reference(ws_met, min_col=2,
                               min_row=perf_data_start,
                               max_row=perf_data_end)
            chart_perf.add_data(vals_p, titles_from_data=True)
            chart_perf.set_categories(cats_p)

            ws_met.add_chart(chart_perf,
                             f'{get_column_letter(len(perf_df.columns) + 2)}{perf_data_start}')

        # Dados completos 120m deste método
        row_det = perf_end + 18 if 'perf_end' in dir() else tbl_end + 20
        row_det = _add_subtitle(ws_met, 'Dados Completos (120m)', row_det)

        cols_export = ['seg_id', 'km_ini', 'km_fim', 'comprimento_m',
                       'Deflexao', 'Deflexao_std', 'Deflexao_caract']
        if col_cluster in seg_final.columns:
            cols_export.append(col_cluster)
        if col_final in seg_final.columns:
            cols_export.append(col_final)
        cols_export = [c for c in cols_export if c in seg_final.columns]

        _write_df(ws_met, seg_final[cols_export].round(4), start_row=row_det)

    # ══════════════════════════════════════════════════════════════════════
    # ABA — AUDITORIA ESTATÍSTICA
    # ══════════════════════════════════════════════════════════════════════
    if df_diag is not None and not df_diag.empty:
        ws_audit = wb.create_sheet('Auditoria')
        ws_audit.sheet_properties.tabColor = 'C00000'

        row = _add_title(ws_audit, 'Auditoria Estatística da Clusterização',
                         merge_cols=len(df_diag.columns))

        # Diagnóstico
        row = _add_subtitle(ws_audit, 'Diagnóstico por Método', row)
        diag_start = row
        row = _write_df(ws_audit, df_diag, start_row=row)
        diag_end = row - 1

        # Colorir parecer
        if 'Parecer' in df_diag.columns:
            parecer_idx = list(df_diag.columns).index('Parecer') + 1
            for r in range(diag_start + 1, diag_end + 1):
                cell = ws_audit.cell(row=r, column=parecer_idx)
                val = str(cell.value) if cell.value else ''
                if '✅' in val:
                    cell.fill = FILL_GREEN
                elif '⚠️' in val or '⚠' in val:
                    cell.fill = FILL_YELLOW
                elif '❌' in val:
                    cell.fill = FILL_RED

        # Gap Statistic
        if gap_df is not None and not gap_df.empty:
            row += 2
            row = _add_subtitle(ws_audit,
                                f'Gap Statistic — k ótimo = {k_otimo}', row)
            gap_data_start = row
            row = _write_df(ws_audit, gap_df.round(4), start_row=row)
            gap_data_end = row - 1

            # Gráfico Gap
            chart_gap = LineChart()
            chart_gap.title = f'Gap Statistic (k ótimo = {k_otimo})'
            chart_gap.style = 10
            chart_gap.width = 18
            chart_gap.height = 10
            chart_gap.x_axis.title = 'k'
            chart_gap.y_axis.title = 'Gap(k)'

            k_col = list(gap_df.columns).index('k') + 1
            gap_col = list(gap_df.columns).index('Gap') + 1

            cats_g = Reference(ws_audit, min_col=k_col,
                               min_row=gap_data_start + 1,
                               max_row=gap_data_end)
            vals_g = Reference(ws_audit, min_col=gap_col,
                               min_row=gap_data_start,
                               max_row=gap_data_end)
            chart_gap.add_data(vals_g, titles_from_data=True)
            chart_gap.set_categories(cats_g)

            ws_audit.add_chart(chart_gap, f'F{gap_data_start}')

        # ARI matrix
        if ari_matrix is not None and not ari_matrix.empty:
            row += 18
            row = _add_subtitle(ws_audit, 'Concordância entre Métodos (ARI)', row)
            ari_export = ari_matrix.reset_index()
            ari_export = ari_export.rename(columns={'index': 'Método'})
            ari_start = row
            row = _write_df(ws_audit, ari_export.round(3), start_row=row)
            ari_end = row - 1

            # Formatação condicional na matriz ARI
            n_met = len(ari_matrix.columns)
            for c in range(2, n_met + 2):
                col_letter = get_column_letter(c)
                ws_audit.conditional_formatting.add(
                    f'{col_letter}{ari_start + 1}:{col_letter}{ari_end}',
                    ColorScaleRule(
                        start_type='num', start_value=0, start_color='FFC7CE',
                        mid_type='num', mid_value=0.5, mid_color='FFEB9C',
                        end_type='num', end_value=1.0, end_color='C6EFCE'
                    )
                )

    # ══════════════════════════════════════════════════════════════════════
    # ABA — ESTATÍSTICAS DESCRITIVAS
    # ══════════════════════════════════════════════════════════════════════
    ws_stats = wb.create_sheet('Estatísticas')
    ws_stats.sheet_properties.tabColor = '7030A0'

    row = _add_title(ws_stats, 'Estatísticas Descritivas', merge_cols=8)

    # Estatísticas gerais dos dados
    desc = seg120[['Deflexao', 'Deflexao_std', 'Deflexao_caract']].describe()
    desc_cols = [c for c in desc.columns if c in seg120.columns]
    if desc_cols:
        desc = desc[desc_cols].round(4)
        row = _add_subtitle(ws_stats, 'Deflexão — Estatísticas Gerais', row)
        desc_reset = desc.reset_index().rename(columns={'index': 'Estatística'})
        row = _write_df(ws_stats, desc_reset, start_row=row)

    # Estatísticas por método — resumo de segmentos
    row += 2
    row = _add_subtitle(ws_stats, 'Resumo por Método', row)

    resumo_metodos = []
    for nome, (seg_f, tbl) in all_downloads.items():
        r = {
            'Método': nome,
            'N_segmentos': len(tbl),
            'Comp_medio_m': tbl['Comprimento_m'].mean() if 'Comprimento_m' in tbl.columns else None,
            'Comp_min_m': tbl['Comprimento_m'].min() if 'Comprimento_m' in tbl.columns else None,
            'Comp_max_m': tbl['Comprimento_m'].max() if 'Comprimento_m' in tbl.columns else None,
            'Defl_media': tbl['Deflexao_med'].mean() if 'Deflexao_med' in tbl.columns else None,
            'Defl_std_media': tbl['Deflexao_std'].mean() if 'Deflexao_std' in tbl.columns else None,
        }
        resumo_metodos.append(r)

    if resumo_metodos:
        df_resumo = pd.DataFrame(resumo_metodos)
        res_start = row
        row = _write_df(ws_stats, df_resumo.round(2), start_row=row)
        res_end = row - 1

        # Gráfico — Número de segmentos por método
        chart_nseg = BarChart()
        chart_nseg.type = 'col'
        chart_nseg.title = 'Número de Segmentos por Método'
        chart_nseg.style = 10
        chart_nseg.width = 18
        chart_nseg.height = 10

        cats_r = Reference(ws_stats, min_col=1, min_row=res_start + 1,
                           max_row=res_end)
        chart_nseg.set_categories(cats_r)

        nseg_col = list(df_resumo.columns).index('N_segmentos') + 1
        data_r = Reference(ws_stats, min_col=nseg_col, min_row=res_start,
                           max_row=res_end)
        chart_nseg.add_data(data_r, titles_from_data=True)
        chart_nseg.series[0].graphicalProperties.solidFill = '7030A0'

        dl_r = DataLabelList()
        dl_r.showVal = True
        chart_nseg.series[0].dLbls = dl_r

        ws_stats.add_chart(chart_nseg, f'A{res_end + 2}')

        # Gráfico — Comprimento médio por método
        if 'Comp_medio_m' in df_resumo.columns:
            chart_comp = BarChart()
            chart_comp.type = 'col'
            chart_comp.title = 'Comprimento Médio dos Segmentos (m)'
            chart_comp.style = 10
            chart_comp.width = 18
            chart_comp.height = 10

            chart_comp.set_categories(cats_r)

            comp_col = list(df_resumo.columns).index('Comp_medio_m') + 1
            data_comp = Reference(ws_stats, min_col=comp_col,
                                  min_row=res_start, max_row=res_end)
            chart_comp.add_data(data_comp, titles_from_data=True)
            chart_comp.series[0].graphicalProperties.solidFill = '548235'

            dl_comp = DataLabelList()
            dl_comp.showVal = True
            dl_comp.numFmt = '#,##0'
            chart_comp.series[0].dLbls = dl_comp

            ws_stats.add_chart(chart_comp, f'J{res_end + 2}')

    # ══════════════════════════════════════════════════════════════════════
    # SALVAR EM BYTES
    # ══════════════════════════════════════════════════════════════════════
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# EXPORTAÇÃO INDIVIDUAL POR MÉTODO — Excel completo standalone
# ══════════════════════════════════════════════════════════════════════════════

def exportar_excel_metodo(
    nome,
    seg120,
    seg_final,
    tbl,
    resultados,
    vars_ok,
    METODO_COL,
    df_comp=None,
    parametros=None,
):
    """
    Gera Excel completo para UM único método, com:

    1. Resumo           — parâmetros, info do método, estatísticas
    2. Segmentos Finais — tabela resumo + gráficos (deflexão, comprimento)
    3. Perfil           — perfil longitudinal com gráfico
    4. Dados 120m       — dados completos a cada 120m
    5. Estatísticas     — descritivas por segmento + histograma
    6. Métricas         — avaliação intrínseca/extrínseca deste método

    Retorna bytes do .xlsx
    """
    wb = Workbook()

    col_cluster = METODO_COL.get(nome, 'cluster_kmeans')
    col_final = f'{col_cluster}_espacial'
    tipo = ('Tradicional' if nome in ['CDA', 'MCV', 'SHS']
            else 'Change-Point' if nome == 'PELT'
            else 'ML Supervisionado' if nome in ['KNN', 'Random Forest']
            else 'Clustering + Linearização')

    # ══════════════════════════════════════════════════════════════════
    # ABA 1 — RESUMO DO MÉTODO
    # ══════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = 'Resumo'
    ws.sheet_properties.tabColor = '1F4E79'

    row = _add_title(ws, f'🛣️ Segmentação — {nome}', merge_cols=8)
    row = _add_subtitle(ws, 'Relatório individual do método', row)

    info = {
        'Data de geração': datetime.now().strftime('%d/%m/%Y %H:%M'),
        'Método': nome,
        'Tipo': tipo,
        'Total de segmentos (120m)': len(seg120),
        'Segmentos finais': len(tbl),
        'Extensão total (m)': f"{seg120['km_fim'].max() - seg120['km_ini'].min():.0f}" if len(seg120) > 0 else 'N/A',
        'Deflexão média geral': f"{seg120['Deflexao'].mean():.2f}" if 'Deflexao' in seg120.columns else 'N/A',
        'Deflexão caract. média': f"{seg120['Deflexao_caract'].mean():.2f}" if 'Deflexao_caract' in seg120.columns else 'N/A',
        'Features usadas': ', '.join(vars_ok),
    }
    if parametros:
        info.update(parametros)

    for key, val in info.items():
        ws.cell(row=row, column=1, value=key).font = Font(bold=True, name='Calibri', size=10)
        ws.cell(row=row, column=1).border = BORDER_THIN
        ws.cell(row=row, column=2, value=str(val)).font = FONT_BODY
        ws.cell(row=row, column=2).border = BORDER_THIN
        ws.cell(row=row, column=2).alignment = ALIGN_LEFT
        row += 1

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 50

    # Métricas intrínsecas neste resumo
    res = resultados.get(nome)
    if res:
        from pipeline.evaluation import avaliar_intrinseco, avaliar_extrinseco_xgb
        metricas = avaliar_intrinseco(res['X_scaled'], res['labels'])
        if metricas:
            row += 1
            row = _add_subtitle(ws, 'Métricas de Qualidade', row)
            for k, v in metricas.items():
                ws.cell(row=row, column=1, value=k).font = Font(bold=True, name='Calibri', size=10)
                ws.cell(row=row, column=1).border = BORDER_THIN
                val_str = f"{v:.4f}" if isinstance(v, float) else str(v)
                ws.cell(row=row, column=2, value=val_str).font = FONT_BODY
                ws.cell(row=row, column=2).border = BORDER_THIN
                row += 1

            extr = avaliar_extrinseco_xgb(res['X_scaled'], res['labels'])
            if extr.get('xgb_accuracy_cv'):
                ws.cell(row=row, column=1, value='XGB Accuracy (CV)').font = Font(bold=True, name='Calibri', size=10)
                ws.cell(row=row, column=1).border = BORDER_THIN
                ws.cell(row=row, column=2, value=f"{extr['xgb_accuracy_cv']:.4f}").font = FONT_BODY
                ws.cell(row=row, column=2).border = BORDER_THIN
                row += 1

    # ══════════════════════════════════════════════════════════════════
    # ABA 2 — SEGMENTOS FINAIS (tabela + gráficos)
    # ══════════════════════════════════════════════════════════════════
    ws_seg = wb.create_sheet('Segmentos Finais')
    ws_seg.sheet_properties.tabColor = '2E75B6'

    row = _add_title(ws_seg, f'Segmentos Finais — {nome}', merge_cols=10)

    # Info rápida
    ws_seg.cell(row=row, column=1, value='Tipo:').font = Font(bold=True, name='Calibri')
    ws_seg.cell(row=row, column=2, value=tipo).font = FONT_BODY
    row += 1
    ws_seg.cell(row=row, column=1, value='Segmentos:').font = Font(bold=True, name='Calibri')
    ws_seg.cell(row=row, column=2, value=len(tbl)).font = FONT_BODY
    row += 1
    if 'Comprimento_m' in tbl.columns:
        ws_seg.cell(row=row, column=1, value='Extensão total (m):').font = Font(bold=True, name='Calibri')
        ws_seg.cell(row=row, column=2, value=f"{tbl['Comprimento_m'].sum():.0f}").font = FONT_BODY
        row += 1
    if 'Deflexao_med' in tbl.columns:
        ws_seg.cell(row=row, column=1, value='Deflexão média (segmentos):').font = Font(bold=True, name='Calibri')
        ws_seg.cell(row=row, column=2, value=f"{tbl['Deflexao_med'].mean():.2f}").font = FONT_BODY
        row += 1
    row += 1

    # Tabela resumo
    row = _add_subtitle(ws_seg, 'Resumo dos Segmentos', row)
    tbl_start = row
    row = _write_df(ws_seg, tbl.round(2), start_row=row)
    tbl_end = row - 1

    # Formatação condicional na deflexão
    if 'Deflexao_med' in tbl.columns:
        defl_idx = list(tbl.columns).index('Deflexao_med') + 1
        defl_letter = get_column_letter(defl_idx)
        ws_seg.conditional_formatting.add(
            f'{defl_letter}{tbl_start + 1}:{defl_letter}{tbl_end}',
            ColorScaleRule(
                start_type='min', start_color='C6EFCE',
                mid_type='percentile', mid_value=50, mid_color='FFEB9C',
                end_type='max', end_color='FFC7CE'
            )
        )

    # Formatação condicional no comprimento
    if 'Comprimento_m' in tbl.columns:
        comp_idx_c = list(tbl.columns).index('Comprimento_m') + 1
        comp_letter = get_column_letter(comp_idx_c)
        ws_seg.conditional_formatting.add(
            f'{comp_letter}{tbl_start + 1}:{comp_letter}{tbl_end}',
            ColorScaleRule(
                start_type='min', start_color='FFC7CE',
                mid_type='percentile', mid_value=50, mid_color='FFEB9C',
                end_type='max', end_color='C6EFCE'
            )
        )

    # Gráfico — Deflexão média por segmento
    if 'Deflexao_med' in tbl.columns and len(tbl) > 0:
        chart1 = BarChart()
        chart1.type = 'col'
        chart1.title = f'Deflexão Média por Segmento — {nome}'
        chart1.style = 10
        chart1.width = 20
        chart1.height = 12
        chart1.y_axis.title = 'Deflexão (0,01mm)'
        chart1.x_axis.title = 'Segmento'

        cats = Reference(ws_seg, min_col=1,
                         min_row=tbl_start + 1, max_row=tbl_end)
        chart1.set_categories(cats)
        data1 = Reference(ws_seg, min_col=defl_idx,
                          min_row=tbl_start, max_row=tbl_end)
        chart1.add_data(data1, titles_from_data=True)
        chart1.series[0].graphicalProperties.solidFill = '2E75B6'

        dl = DataLabelList()
        dl.showVal = True
        dl.numFmt = '0.00'
        chart1.series[0].dLbls = dl
        ws_seg.add_chart(chart1, f'A{tbl_end + 2}')

    # Gráfico — Deflexão caract por segmento
    if 'Deflexao_caract_med' in tbl.columns and len(tbl) > 0:
        dc_idx = list(tbl.columns).index('Deflexao_caract_med') + 1
        chart_dc = BarChart()
        chart_dc.type = 'col'
        chart_dc.title = f'Deflexão Característica Média — {nome}'
        chart_dc.style = 10
        chart_dc.width = 20
        chart_dc.height = 12
        chart_dc.y_axis.title = 'Deflexão Caract. (0,01mm)'

        chart_dc.set_categories(cats)
        data_dc = Reference(ws_seg, min_col=dc_idx,
                            min_row=tbl_start, max_row=tbl_end)
        chart_dc.add_data(data_dc, titles_from_data=True)
        chart_dc.series[0].graphicalProperties.solidFill = 'C00000'

        dl_dc = DataLabelList()
        dl_dc.showVal = True
        dl_dc.numFmt = '0.00'
        chart_dc.series[0].dLbls = dl_dc
        ws_seg.add_chart(chart_dc, f'L{tbl_end + 2}')

    # Gráfico — Comprimento dos segmentos
    if 'Comprimento_m' in tbl.columns and len(tbl) > 0:
        chart2 = BarChart()
        chart2.type = 'col'
        chart2.title = f'Comprimento dos Segmentos — {nome}'
        chart2.style = 10
        chart2.width = 20
        chart2.height = 12
        chart2.y_axis.title = 'Comprimento (m)'
        chart2.x_axis.title = 'Segmento'

        chart2.set_categories(cats)
        data2 = Reference(ws_seg, min_col=comp_idx_c,
                          min_row=tbl_start, max_row=tbl_end)
        chart2.add_data(data2, titles_from_data=True)
        chart2.series[0].graphicalProperties.solidFill = '548235'

        dl2 = DataLabelList()
        dl2.showVal = True
        dl2.numFmt = '#,##0'
        chart2.series[0].dLbls = dl2
        ws_seg.add_chart(chart2, f'A{tbl_end + 20}')

    # Gráfico — Deflexão std por segmento
    if 'Deflexao_std' in tbl.columns and len(tbl) > 0:
        std_idx = list(tbl.columns).index('Deflexao_std') + 1
        chart_std = BarChart()
        chart_std.type = 'col'
        chart_std.title = f'Desvio Padrão da Deflexão — {nome}'
        chart_std.style = 10
        chart_std.width = 20
        chart_std.height = 12
        chart_std.y_axis.title = 'Std (0,01mm)'

        chart_std.set_categories(cats)
        data_std = Reference(ws_seg, min_col=std_idx,
                             min_row=tbl_start, max_row=tbl_end)
        chart_std.add_data(data_std, titles_from_data=True)
        chart_std.series[0].graphicalProperties.solidFill = 'BF8F00'

        dl_std = DataLabelList()
        dl_std.showVal = True
        dl_std.numFmt = '0.00'
        chart_std.series[0].dLbls = dl_std
        ws_seg.add_chart(chart_std, f'L{tbl_end + 20}')

    # ══════════════════════════════════════════════════════════════════
    # ABA 3 — PERFIL LONGITUDINAL
    # ══════════════════════════════════════════════════════════════════
    ws_perf = wb.create_sheet('Perfil Longitudinal')
    ws_perf.sheet_properties.tabColor = '548235'

    row = _add_title(ws_perf, f'Perfil Longitudinal — {nome}', merge_cols=8)

    if col_final in seg_final.columns:
        perf_cols = ['km_ini', 'Deflexao']
        if 'Deflexao_caract' in seg_final.columns:
            perf_cols.append('Deflexao_caract')
        if 'Deflexao_std' in seg_final.columns:
            perf_cols.append('Deflexao_std')
        perf_cols.append(col_final)

        perf_cols = [c for c in perf_cols if c in seg_final.columns]
        perf_df = seg_final[perf_cols].copy()
        perf_df = perf_df.rename(columns={col_final: 'Segmento'})

        perf_data_start = row
        row = _write_df(ws_perf, perf_df.round(2), start_row=row)
        perf_data_end = row - 1

        # Gráfico de perfil — Deflexão
        chart_p = LineChart()
        chart_p.title = f'Perfil Longitudinal de Deflexão — {nome}'
        chart_p.style = 10
        chart_p.width = 26
        chart_p.height = 14
        chart_p.x_axis.title = 'Estação (m)'
        chart_p.y_axis.title = 'Deflexão (0,01mm)'

        cats_p = Reference(ws_perf, min_col=1,
                           min_row=perf_data_start + 1, max_row=perf_data_end)
        vals_p = Reference(ws_perf, min_col=2,
                           min_row=perf_data_start, max_row=perf_data_end)
        chart_p.add_data(vals_p, titles_from_data=True)
        chart_p.set_categories(cats_p)
        chart_p.series[0].graphicalProperties.line.width = 15000

        # Deflexão caract no mesmo gráfico
        if 'Deflexao_caract' in perf_df.columns:
            dc_p_idx = list(perf_df.columns).index('Deflexao_caract') + 1
            vals_dc = Reference(ws_perf, min_col=dc_p_idx,
                                min_row=perf_data_start, max_row=perf_data_end)
            chart_p.add_data(vals_dc, titles_from_data=True)

        ws_perf.add_chart(chart_p,
                          f'{get_column_letter(len(perf_df.columns) + 2)}{perf_data_start}')

    # ══════════════════════════════════════════════════════════════════
    # ABA 4 — DADOS COMPLETOS 120m
    # ══════════════════════════════════════════════════════════════════
    ws_dados = wb.create_sheet('Dados 120m')
    ws_dados.sheet_properties.tabColor = '7030A0'

    row = _add_title(ws_dados, f'Dados Completos (120m) — {nome}',
                     merge_cols=10)

    cols_export = ['seg_id', 'km_ini', 'km_fim', 'comprimento_m',
                   'Deflexao', 'Deflexao_std', 'Deflexao_caract',
                   'Deflexao_max', 'Deflexao_min', 'CV_deflexao']
    if col_cluster in seg_final.columns:
        cols_export.append(col_cluster)
    if col_final in seg_final.columns:
        cols_export.append(col_final)
    cols_export = [c for c in cols_export if c in seg_final.columns]

    data120_start = row
    row = _write_df(ws_dados, seg_final[cols_export].round(4), start_row=row)
    data120_end = row - 1

    # Formatação condicional na deflexão
    if 'Deflexao' in cols_export:
        d_idx = cols_export.index('Deflexao') + 1
        d_letter = get_column_letter(d_idx)
        ws_dados.conditional_formatting.add(
            f'{d_letter}{data120_start + 1}:{d_letter}{data120_end}',
            ColorScaleRule(
                start_type='min', start_color='C6EFCE',
                mid_type='percentile', mid_value=50, mid_color='FFEB9C',
                end_type='max', end_color='FFC7CE'
            )
        )

    # Gráfico de perfil nos dados 120m
    if 'km_ini' in cols_export and 'Deflexao' in cols_export:
        km_idx = cols_export.index('km_ini') + 1
        d_idx = cols_export.index('Deflexao') + 1

        chart_d = LineChart()
        chart_d.title = f'Deflexão a cada 120m — {nome}'
        chart_d.style = 10
        chart_d.width = 24
        chart_d.height = 12
        chart_d.x_axis.title = 'Estação (m)'
        chart_d.y_axis.title = 'Deflexão (0,01mm)'

        cats_d = Reference(ws_dados, min_col=km_idx,
                           min_row=data120_start + 1, max_row=data120_end)
        vals_d = Reference(ws_dados, min_col=d_idx,
                           min_row=data120_start, max_row=data120_end)
        chart_d.add_data(vals_d, titles_from_data=True)
        chart_d.set_categories(cats_d)

        ws_dados.add_chart(chart_d,
                           f'{get_column_letter(len(cols_export) + 2)}{data120_start}')

    # ══════════════════════════════════════════════════════════════════
    # ABA 5 — ESTATÍSTICAS POR SEGMENTO
    # ══════════════════════════════════════════════════════════════════
    ws_stats = wb.create_sheet('Estatísticas')
    ws_stats.sheet_properties.tabColor = 'BF8F00'

    row = _add_title(ws_stats, f'Estatísticas — {nome}', merge_cols=8)

    # Descritivas gerais
    row = _add_subtitle(ws_stats, 'Estatísticas Gerais da Deflexão', row)
    desc_cols = [c for c in ['Deflexao', 'Deflexao_std', 'Deflexao_caract',
                             'CV_deflexao'] if c in seg120.columns]
    if desc_cols:
        desc = seg120[desc_cols].describe().round(4)
        desc_reset = desc.reset_index().rename(columns={'index': 'Estatística'})
        row = _write_df(ws_stats, desc_reset, start_row=row)

    # Descritivas por segmento
    row += 2
    row = _add_subtitle(ws_stats, 'Estatísticas por Segmento Final', row)

    if col_final in seg_final.columns:
        stats_per_seg = seg_final.groupby(col_final).agg(
            N_pontos_120m=('seg_id', 'count'),
            Deflexao_media=('Deflexao', 'mean'),
            Deflexao_std=('Deflexao', 'std'),
            Deflexao_min=('Deflexao', 'min'),
            Deflexao_max=('Deflexao', 'max'),
            Deflexao_mediana=('Deflexao', 'median'),
            CV=('Deflexao', lambda x: x.std() / x.mean() if x.mean() != 0 else 0),
        ).reset_index().rename(columns={col_final: 'Segmento'})

        stats_start = row
        row = _write_df(ws_stats, stats_per_seg.round(4), start_row=row)
        stats_end = row - 1

        # Formatação condicional no CV (alto = ruim)
        if 'CV' in stats_per_seg.columns:
            cv_idx = list(stats_per_seg.columns).index('CV') + 1
            cv_letter = get_column_letter(cv_idx)
            ws_stats.conditional_formatting.add(
                f'{cv_letter}{stats_start + 1}:{cv_letter}{stats_end}',
                ColorScaleRule(
                    start_type='min', start_color='C6EFCE',
                    mid_type='percentile', mid_value=50, mid_color='FFEB9C',
                    end_type='max', end_color='FFC7CE'
                )
            )

        # Gráfico — Deflexão média + std por segmento
        if len(stats_per_seg) > 0:
            chart_s = BarChart()
            chart_s.type = 'col'
            chart_s.title = f'Deflexão por Segmento (média ± std) — {nome}'
            chart_s.style = 10
            chart_s.width = 22
            chart_s.height = 12

            cats_s = Reference(ws_stats, min_col=1,
                               min_row=stats_start + 1, max_row=stats_end)
            chart_s.set_categories(cats_s)

            med_idx = list(stats_per_seg.columns).index('Deflexao_media') + 1
            data_m = Reference(ws_stats, min_col=med_idx,
                               min_row=stats_start, max_row=stats_end)
            chart_s.add_data(data_m, titles_from_data=True)
            chart_s.series[0].graphicalProperties.solidFill = '2E75B6'

            std_s_idx = list(stats_per_seg.columns).index('Deflexao_std') + 1
            data_st = Reference(ws_stats, min_col=std_s_idx,
                                min_row=stats_start, max_row=stats_end)
            chart_s.add_data(data_st, titles_from_data=True)
            chart_s.series[1].graphicalProperties.solidFill = 'BF8F00'

            ws_stats.add_chart(chart_s, f'A{stats_end + 2}')

            # Gráfico — CV por segmento
            chart_cv = BarChart()
            chart_cv.type = 'col'
            chart_cv.title = f'Coeficiente de Variação por Segmento — {nome}'
            chart_cv.style = 10
            chart_cv.width = 22
            chart_cv.height = 12

            chart_cv.set_categories(cats_s)
            data_cv = Reference(ws_stats, min_col=cv_idx,
                                min_row=stats_start, max_row=stats_end)
            chart_cv.add_data(data_cv, titles_from_data=True)
            chart_cv.series[0].graphicalProperties.solidFill = 'C00000'

            dl_cv = DataLabelList()
            dl_cv.showVal = True
            dl_cv.numFmt = '0.00'
            chart_cv.series[0].dLbls = dl_cv

            ws_stats.add_chart(chart_cv, f'L{stats_end + 2}')

    # ══════════════════════════════════════════════════════════════════
    # ABA 6 — POSIÇÃO NESTE MÉTODO VS OUTROS (se df_comp disponível)
    # ══════════════════════════════════════════════════════════════════
    if df_comp is not None and not df_comp.empty:
        ws_rank = wb.create_sheet('Ranking')
        ws_rank.sheet_properties.tabColor = 'C00000'

        row = _add_title(ws_rank, f'Posição de {nome} entre todos os métodos',
                         merge_cols=len(df_comp.columns))

        row = _add_subtitle(ws_rank, 'Tabela Comparativa (todos os métodos)', row)
        comp_start = row
        row = _write_df(ws_rank, df_comp.round(4), start_row=row)
        comp_end = row - 1

        # Highlight a linha deste método
        metodo_col_name = df_comp.columns[0]  # geralmente 'Método'
        for r in range(comp_start + 1, comp_end + 1):
            cell_val = ws_rank.cell(row=r, column=1).value
            if cell_val == nome:
                for c in range(1, len(df_comp.columns) + 1):
                    ws_rank.cell(row=r, column=c).fill = PatternFill(
                        start_color='FFFF00', end_color='FFFF00',
                        fill_type='solid')
                    ws_rank.cell(row=r, column=c).font = Font(
                        bold=True, name='Calibri', size=10)

        # Gráfico comparativo
        if 'silhouette' in df_comp.columns:
            sil_col = list(df_comp.columns).index('silhouette') + 1
            chart_r = BarChart()
            chart_r.type = 'col'
            chart_r.title = 'Silhouette Score — Todos os Métodos'
            chart_r.style = 10
            chart_r.width = 20
            chart_r.height = 12

            cats_r = Reference(ws_rank, min_col=1,
                               min_row=comp_start + 1, max_row=comp_end)
            chart_r.set_categories(cats_r)
            data_r = Reference(ws_rank, min_col=sil_col,
                               min_row=comp_start, max_row=comp_end)
            chart_r.add_data(data_r, titles_from_data=True)
            chart_r.series[0].graphicalProperties.solidFill = '2E75B6'

            dl_r = DataLabelList()
            dl_r.showVal = True
            dl_r.numFmt = '0.000'
            chart_r.series[0].dLbls = dl_r
            ws_rank.add_chart(chart_r, f'A{comp_end + 2}')

        if 'davies_bouldin' in df_comp.columns:
            db_col = list(df_comp.columns).index('davies_bouldin') + 1
            chart_db = BarChart()
            chart_db.type = 'col'
            chart_db.title = 'Davies-Bouldin — Todos os Métodos (menor=melhor)'
            chart_db.style = 10
            chart_db.width = 20
            chart_db.height = 12

            chart_db.set_categories(cats_r)
            data_db = Reference(ws_rank, min_col=db_col,
                                min_row=comp_start, max_row=comp_end)
            chart_db.add_data(data_db, titles_from_data=True)
            chart_db.series[0].graphicalProperties.solidFill = 'BF8F00'
            ws_rank.add_chart(chart_db, f'L{comp_end + 2}')

    # ══════════════════════════════════════════════════════════════════
    # SALVAR
    # ══════════════════════════════════════════════════════════════════
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
